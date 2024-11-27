import time  # Importa a biblioteca time para manipulação de tempo
from selenium import webdriver  # Importa o webdriver do Selenium para automação de navegadores
from selenium.webdriver.common.keys import Keys  # Importa Keys do Selenium para simular pressionamento de teclas
from selenium.webdriver.common.by import By  # Importa By do Selenium para localizar elementos na página
from openpyxl import load_workbook  # Importa load_workbook do openpyxl para manipulação de arquivos Excel
from datetime import datetime  # Importa datetime para manipulação de datas e horas
from tkinter import *  # Importa o módulo tkinter para criar interfaces gráficas


class HistoricoTemperatura:
    # Inicializando a classe com o arquivo Excel e a planilha
    def __init__(self):
        self.arquivo_excel = 'historico_temperatura.xlsx'
        self.planilha_nome = 'temperatura'

    # Inicializando o navegador Chrome, abrindo o Google e pesquisando a temperatura atual
    def iniciar_navegador(self):
        self.navegador = webdriver.Chrome()  # Inicia o navegador
        self.navegador.get('https://www.google.com/')  # Abre o Google
        time.sleep(1)
        
        # Localizando o campo de pesquisa e inserindo o texto
        campo_pesquisa = self.navegador.find_element(By.XPATH, '//*[@id="APjFqb"]')
        campo_pesquisa.send_keys('temperatura atual')
        time.sleep(1)
        print("Pesquisando temperatura atual...")
        # Localiza o boao de pesquisa e clica no botão "Enter"
        campo_pesquisa.send_keys(Keys.ENTER)
        time.sleep(3)

    # Carregar o arquivo Excel e a planilha
    def carregar_planilha(self):
        self.workbook = load_workbook(self.arquivo_excel)
        self.planilha = self.workbook[self.planilha_nome]
    
    
    # Obter a próxima linha vazia
    def proxima_linha(self):
        return self.planilha.max_row + 1
    
    # Obter a data e hora atual para preencher na planilha
    def obter_data_hora_atual(self):
        self.data_hora_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"Data e hora atual: {self.data_hora_atual}")
    
    # Obter a temperatura atual do Google
    def obter_temperatura_atual(self):
        # Pega o conteúdo do innerHTML do elemento
        self.temperatura = self.navegador.find_element(By.XPATH, '//*[@id="wob_tm"]').get_attribute('innerHTML')
        try:
            self.temperatura = int(self.temperatura)
        except:
            self.temperatura = None
        print(f"Temperatura atual: {self.temperatura}°C")
        
    # Obter o alerta de umidade
    def obter_alerta_umidade(self):
        try:
            self.alerta_umidade = self.navegador.find_element(By.XPATH, '//*[@id="wob_wc"]/div[4]/div[1]').get_attribute('innerHTML')
            print(f"Alerta de umidade: {self.alerta_umidade}")
        except:
            self.alerta_umidade = 'Alerta Umidade nao encontrado'
            print("Alerta de umidade não encontrado")
    
    # Salvar os dados na planilha
    def salvar_dados(self):
        
        linha_atual = self.proxima_linha()
        self.planilha.cell(row=linha_atual, column=1).value = self.data_hora_atual
        self.planilha.cell(row=linha_atual, column=2).value = self.temperatura
        self.planilha.cell(row=linha_atual, column=3).value = self.alerta_umidade
        
        print(f"Dados salvos com sucesso na linha {linha_atual}")
    
    # Salvar o arquivo Excel
    def salvar_arquivo(self):
        # Salvar o arquivo Excel
        self.workbook.save(self.arquivo_excel)
        print(f"Arquivo {self.arquivo_excel} salvo com sucesso")

        
    def executar_programa(self):
        # Iniciar o navegador
        self.iniciar_navegador()
        # Carregar a planilha
        self.carregar_planilha()
        # Salvar a data e hora na próxima linha disponível
        self.obter_data_hora_atual()
        # Obter a temperatura atual
        self.obter_temperatura_atual()
        # Obter o alerta de umidade
        self.obter_alerta_umidade()
        # Salvar os dados na planilha
        self.salvar_dados()
        # Salvar o arquivo Excel
        self.salvar_arquivo()
        
class GerarRelatoriosApp:
    def __init__(self):
        self.layout = Tk() # Inicializa a janela
        self.layout.title('Buscar previsão do tempo') # Define o título da janela
        self.layout.geometry('300x200') # Define o tamanho da janela
        self.tela = Frame(self.layout) # Cria um frame na janela
        
        self.label = Label(self.tela, text='Clique no Botao pra pegar a Temperatura') # Cria um label no frame
        self.botao = Button(self.tela, text='Gerar Relatório', command=self.gerar_relatorio) # Cria um botao no frame
        
        self.tela.pack() # Empacota o frame na janela
        self.botao.pack() # Empacota o botao no frame
        
        mainloop() # Inicia o loop principal da janela
        
    # Função para gerar o relatório
    def gerar_relatorio(self):
        try:
            historico = HistoricoTemperatura()
            historico.executar_programa()
        except Exception as e:
            print(f"Erro ao gerar relatório: {e}")
        
        
# Executar o programa
if __name__ == "__main__":
    app = GerarRelatoriosApp()